import streamlit as st
import pandas as pd
from datetime import date
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import os

st.set_page_config(page_title="Risk Assessment Form", layout="centered")

# ✅ Simple password protection
correct_password = "engineer2024"
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    password_input = st.text_input("Enter password to access the app", type="password")
    if password_input == correct_password:
        st.session_state.authenticated = True
        st.rerun()
    else:
        st.stop()

# ✅ Define logo path
logo_path = r"C:\Users\davaa\Desktop\logo.jpg"

# ✅ Display logo in the UI
logo = Image.open(logo_path)
st.image(logo, width=300)

# Risk matrix
risk_matrix = pd.DataFrame([
    ["Extensive", "Certain", "High"],
    ["Extensive", "Probable", "High"],
    ["Extensive", "Likely", "High"],
    ["Extensive", "Unlikely", "High"],
    ["Extensive", "Rare", "Medium"],

    ["Major", "Certain", "High"],
    ["Major", "Probable", "High"],
    ["Major", "Likely", "High"],
    ["Major", "Unlikely", "Medium"],
    ["Major", "Rare", "Medium"],

    ["Significant", "Certain", "High"],
    ["Significant", "Probable", "High"],
    ["Significant", "Likely", "Medium"],
    ["Significant", "Unlikely", "Medium"],
    ["Significant", "Rare", "Low"],

    ["Minor", "Certain", "Medium"],
    ["Minor", "Probable", "Medium"],
    ["Minor", "Likely", "Low"],
    ["Minor", "Unlikely", "Low"],
    ["Minor", "Rare", "Low"],

    ["Negligible", "Certain", "Low"],
    ["Negligible", "Probable", "Low"],
    ["Negligible", "Likely", "Low"],
    ["Negligible", "Unlikely", "Low"],
    ["Negligible", "Rare", "Low"]
], columns=["Consequence Rating", "Likelihood", "RiskRating"])

# State management
if "partial_entry" not in st.session_state:
    st.session_state.partial_entry = None
if "entries" not in st.session_state:
    st.session_state.entries = []

st.title("✍️ Risk Assessment Form")

# Step 1: Hazard input
with st.form("step1_form"):
    st.subheader("Step 1: Define Hazard and Calculate Risk")
    hazard = st.text_input("Hazard description")
    consequences = st.text_input("Consequences")
    consequence_rating = st.selectbox("Consequence Rating", ["Extensive", "Major", "Significant", "Minor", "Negligible"])
    likelihood = st.selectbox("Likelihood", ["Certain", "Probable", "Likely", "Unlikely", "Rare"])
    step1_submitted = st.form_submit_button("Add Entry")

    if step1_submitted:
        if not hazard or not consequences:
            st.warning("Please enter both Hazard and Consequences.")
        else:
            match = risk_matrix[
                (risk_matrix["Consequence Rating"] == consequence_rating) &
                (risk_matrix["Likelihood"] == likelihood)
            ]
            risk_rating = match["RiskRating"].values[0] if not match.empty else "Unknown"
            st.session_state.partial_entry = {
                "Hazard": hazard,
                "Date": date.today(),
                "Consequences": consequences,
                "Consequence Rating": consequence_rating,
                "Likelihood": likelihood,
                "RiskRating": risk_rating
            }

# Step 2: Action input
if st.session_state.partial_entry:
    partial = st.session_state.partial_entry
    st.info(f"📊 **Calculated Risk Rating**: {partial['RiskRating']}")
    with st.form("step2_form"):
        st.subheader("Step 2: Enter Recommended Actions")
        actions = st.text_area("Recommended Actions")
        step2_submitted = st.form_submit_button("Save to Table")

        if step2_submitted:
            if not actions:
                st.warning("Please provide a recommended action.")
            else:
                full_entry = {**partial, "Actions": actions}
                st.session_state.entries.append(full_entry)
                st.session_state.partial_entry = None
                st.success("Entry added to Risk Assessment Table.")

# Display and style the table
if st.session_state.entries:
    st.subheader("📋 Risk Assessment Table")
    df = pd.DataFrame(st.session_state.entries)
    expected_cols = ["Hazard", "Date", "Consequences", "Consequence Rating", "Likelihood", "RiskRating", "Actions"]
    df = df.reindex(columns=expected_cols)

    def highlight_risk(val):
        if val == "High":
            return "background-color: #ffcccc; color: red; font-weight: bold"
        elif val == "Medium":
            return "background-color: #fff8b0; color: #b36b00; font-weight: bold"
        elif val == "Low":
            return "background-color: #d6f5d6; color: green; font-weight: bold"
        return ""

    styled_df = df.style.map(highlight_risk, subset=["RiskRating"])
    st.dataframe(styled_df, use_container_width=True)

    # Export buttons remain unchanged (using updated column names)


    # Excel Export
    if st.button("📥 Export Table to Excel"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Risk Assessment"

        logo_img = XLImage(logo_path)
        logo_img.width = 200
        logo_img.height = 60
        ws.add_image(logo_img, "A1")

        fills = {
            "High": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
            "Medium": PatternFill(start_color="FFF8B0", end_color="FFF8B0", fill_type="solid"),
            "Low": PatternFill(start_color="D6F5D6", end_color="D6F5D6", fill_type="solid")
        }

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=5, column=col_idx, value=col_name)
            cell.font = Font(bold=True)

        for row_idx, row in df.iterrows():
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=row_idx + 6, column=col_idx, value=row[col_name])
                if col_name == "RiskRating" and row[col_name] in fills:
                    cell.fill = fills[row[col_name]]

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        st.download_button(
            label="📄 Download Excel File",
            data=excel_buffer,
            file_name=f"risk_assessment_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


